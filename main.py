# conding=utf-8
import tkinter as tk
import tkinter.filedialog
import os
import glob

import pandas as pd
import openpyxl

from openpyxl.worksheet.header_footer import _HeaderFooterPart


window = tk.Tk()
window.title("auto excel")
window.geometry('450x360')
var1 = tk.StringVar()
fileName = tk.Entry(window, textvariable=var1, width=40).place(x=100, y=20)

statusStr = tk.StringVar()    # 将label标签的内容设置为字符类型，用var来接收函数的传出内容用以显示在标签上
l = tk.Label(window, textvariable=statusStr, bg='green',
             fg='white', font=('Arial', 12), width=30, height=2)
l.pack(side='bottom')


def chooseFile():
    fn1 = tk.filedialog.askdirectory()
    var1.set(fn1)
    statusStr.set('')





tk.Button(window, text='选择文件夹', command=chooseFile).place(x=10, y=20)

# tk.Label(window, text='固定的行数').place(x=10, y=80)
# var2 = tk.StringVar()
# var2.set('4')
# tk.Entry(window, textvariable=var2, width=40).place(x=100, y=80)




def convertCore(filePath, name):
    wb = openpyxl.load_workbook(filePath)
    # fixColNum = var2.get()
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        firstSheetName=wb.sheetnames[0]
        if sheetname != firstSheetName:
            dataFrame = pd.read_excel(filePath, sheet_name=sheetname)
            dataLen=len(dataFrame)+1
            printArea='A1:T'+str(dataLen)
            ws.print_area = printArea

            ws.print_title_rows = '1:4'
            ws.row_dimensions[4].height = 36
            ws.HeaderFooter.differentFirst = False
            ws.HeaderFooter.differentOddEven = True
            ws.oddHeader.right.text = ''
            ws.oddHeader.center.text = ''
            ws.oddHeader.left.text = ''
            ws.oddFooter.left = _HeaderFooterPart(
                '制表：曹开升                            校核：高  进                             审核：吴贵年                            项目负责人：楼少华')

            ws.oddFooter.right.size = 8
            ws.oddFooter.left.size = 8
            ws.oddFooter.center.size = 8
            ws.oddFooter.right.font = "宋体"
            ws.oddFooter.right.text = "第 &[Page] 页"
            ws.oddFooter.center.text = ""
            ws.evenFooter.left = _HeaderFooterPart(
                '制表：曹开升                            校核：高  进                             审核：吴贵年                            项目负责人：楼少华')
            ws.evenFooter.right.size = 8
            ws.evenFooter.left.size = 8
            ws.evenFooter.right.font = "宋体"
            ws.evenFooter.right.text = "第 &[Page] 页"
            ws.evenFooter.center.text = ""

            # ws.page_setup.fitToHeight = True

            ws.page_margins = openpyxl.worksheet.page.PageMargins(
                left=0.1, right=0.1, top=0.5, bottom=0.6, header=0.5, footer=0.5)

            # ws.merge_cells("G2:N2")

            ws.cell(2, 7).value = name


        else:
            ws.cell(4, 1).value = name
            ws.cell(12, 1).value = name
            printArea='A1:Q22'
            ws.row_dimensions[9].height=22.5
            ws.print_area = printArea
                

    wb.save(filePath)
    


def bianLi(rootDir):
    for root, dirs, files in os.walk(rootDir):
        for file in files:
            filePath = os.path.join(root, file)
            if filePath.endswith('.xlsx'):
                tempName = os.path.splitext(file)[0]
                convertCore(filePath, tempName)
        for dir in dirs:
            bianLi(dir)


def batchConvert():
    filePath = var1.get()
    # myPathRoot(filePath)
    if filePath == "":
        statusStr.set('请选择文件夹！！')
        return

    xlsx_file_number = glob.glob(pathname=filePath+'/' + r'*.xlsx')


    if len(xlsx_file_number) == 0:
        statusStr.set('不存在.xlsx的文件')
        return
    bianLi(filePath)
    statusStr.set('恭喜，转换完毕！')


tk.Button(window, text='开始', width=50,
          command=batchConvert).place(x=10, y=260)
window.mainloop()


import os
import openpyxl
from openpyxl.worksheet.header_footer import _HeaderFooterPart
import pandas as pd
from openpyxl.utils import get_column_letter

from openpyxl.drawing.image import Image

from openpyxl.styles import Color, Font, Alignment
xlsxFiles = (fn for fn in os.listdir('.') if fn.endswith('.xlsx'))

alignObj1 = Alignment(horizontal='center',
                      vertical='center', wrapText=True)
fontObj1 = Font(name=u'楷体_GB2312', bold=True, italic=False, size=24)
fontObj2 = Font(name=u'楷体_GB2312', bold=True, italic=False, size=30)
fontObj3 = Font(name=u'楷体', bold=True, italic=False, size=18)
fontObj4 = Font(name=u'宋体', bold=True, italic=False, size=26)
fontObj5 = Font(name=u'宋体', bold=False, italic=False, size=15)


def constructorExcel(ws, startNum, value1, value2, value3):
    ws['A'+str(startNum)].value = value1
    ws['A'+str(startNum)].font = fontObj5
    ws['B'+str(startNum)].value = value2
    ws['B'+str(startNum)].font = fontObj5
    ws['C'+str(startNum)].value = value3
    ws['C'+str(startNum)].font = fontObj5
    ws.row_dimensions[startNum].height = 38.45


for xlsxFile in xlsxFiles:
    wb = openpyxl.load_workbook(xlsxFile)

    for ws in wb.worksheets:
        fillValue = ws.cell(2, 1).value
        if ws.cell(2, 1).value != "":
            if '管线类型：' in fillValue:
                maxColumnLetter = get_column_letter(ws.max_column)
                mergeRange = "A2:"+maxColumnLetter+"2"
                ws.merge_cells(mergeRange)
                ws.cell(2, 1).value = fillValue

        ws.HeaderFooter.differentOddEven = True
        ws.oddFooter.left = _HeaderFooterPart('探测单位：中国电建集团华东勘测设计研究院有限公司')
        ws.oddFooter.center = _HeaderFooterPart('制表者：段磊仔    校核者：陈皎')

        ws.oddFooter.right.size = 8
        ws.oddFooter.left.size = 8
        ws.oddFooter.center.size = 8
        ws.oddFooter.right.font = "宋体"
        ws.oddFooter.right.text = "日期:2019-5      第 &[Page]-2 页    共 &N-2 页"
        ws.evenFooter.left = _HeaderFooterPart('探测单位：中国电建集团华东勘测设计研究院有限公司')
        ws.evenFooter.center = _HeaderFooterPart('制表者：段磊仔    校核者：陈皎')
        ws.evenFooter.right.size = 8
        ws.evenFooter.left.size = 8
        ws.evenFooter.center.size = 8
        ws.evenFooter.right.font = "宋体"
        ws.evenFooter.right.text = "日期:2019-5      第 &[Page]-2 页    共 &N-2 页"

    ws0 = wb['目录']
    dataFrame = pd.read_excel(xlsxFile, sheet_name='目录')
    dataDict = dataFrame.to_dict(orient='list')
    wb.remove(wb['目录'])
    ws00 = wb.create_sheet("目录", 0)
    ws00.row_dimensions[1].height = 15.6
    ws00.row_dimensions[2].height = 50.1
    ws00.row_dimensions[3].height = 51

    ws00.merge_cells("A2:C2")
    ws00.cell(2, 1).value = "目     录"
    ws00.cell(2, 1).alignment = alignObj1
    ws00.cell(2, 1).font = fontObj4
    ws00.column_dimensions['A'].width = 11.38
    ws00.column_dimensions['B'].width = 90.36
    ws00.column_dimensions['C'].width = 9.75

    splitStr = ' ………………………………………………… '
    fillStr = '………………………………………………………………………………………………………………'
    startNum = 4
    for i in dataDict['管 线 点 成 果 表']:
        if pd.isnull(i) == False:
            if splitStr in str(i):
                itemValue = i.split(splitStr)
                constructorExcel(
                    ws00, startNum, itemValue[0], fillStr, itemValue[1])
                startNum = startNum+1

    ws1 = wb.create_sheet("封面", 0)

    # 调整列宽
    ws1.column_dimensions['A'].width = 117.88

    # 调整行高
    for i in range(1, 9):
        ws1.row_dimensions[i].height = 60

    ws1['A2'] = "市政文锦渠及东湖公园暗涵综合整治和清污剥离工程—文锦渡口岸泵站"
    ws1['A2'].font = fontObj1
    ws1['A2'].alignment = alignObj1
    ws1['A3'] = "管线点成果表"
    ws1['A3'].font = fontObj2
    ws1['A3'].alignment = alignObj1
    ws1['A8'] = "二〇一九年九月"
    ws1['A8'].font = fontObj3
    ws1['A8'].alignment = alignObj1

    img = Image('images/img.png')
    newsize = (567, 71)
    img.width, img.height = newsize  # 这两个属性分别是对应添加图片的宽高

    img.anchor = 'A7'

    ws1.add_image(img)
    wb.save('new_'+xlsxFile)

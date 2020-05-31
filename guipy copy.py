# conding=utf-8
import tkinter as tk
import tkinter.filedialog
import os
import glob

import pandas as pd
import openpyxl
from openpyxl.worksheet.header_footer import _HeaderFooterPart
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.styles import Color, Font, Alignment


alignObj1 = Alignment(horizontal='center',
                      vertical='center', wrapText=True)
fontObj1 = Font(name=u'楷体_GB2312', bold=True, italic=False, size=24)
fontObj2 = Font(name=u'楷体_GB2312', bold=True, italic=False, size=30)
fontObj3 = Font(name=u'楷体', bold=True, italic=False, size=18)
fontObj4 = Font(name=u'宋体', bold=True, italic=False, size=26)
fontObj5 = Font(name=u'宋体', bold=False, italic=False, size=15)


window = tk.Tk()
window.title("批量脚注")
window.geometry('450x360')
var1 = tk.StringVar()
fileName = tk.Entry(window, textvariable=var1, width=40).place(x=100, y=20)

statusStr = tk.StringVar()    # 将label标签的内容设置为字符类型，用var来接收函数的传出内容用以显示在标签上
l = tk.Label(window, textvariable=statusStr, bg='green',
             fg='white', font=('Arial', 12), width=30, height=2)
l.pack(side='bottom')


def chooseFile():
    fn1 = tk.filedialog.askdirectory()
    var1.set(fn1)
    statusStr.set('')


def constructorExcel(ws, startNum, value1, value2, value3):
    ws['A'+str(startNum)].value = value1
    ws['A'+str(startNum)].font = fontObj5
    ws['B'+str(startNum)].value = value2
    ws['B'+str(startNum)].font = fontObj5
    ws['C'+str(startNum)].value = value3
    ws['C'+str(startNum)].font = fontObj5
    ws.row_dimensions[startNum].height = 38.45


tk.Button(window, text='选择文件夹', command=chooseFile).place(x=10, y=20)

tk.Label(window, text='页脚左边的文字').place(x=10, y=80)
var2 = tk.StringVar()
var2.set('左边的文字')
tk.Entry(window, textvariable=var2, width=40).place(x=100, y=80)

tk.Label(window, text='页脚中间的文字').place(x=10, y=140)
var3 = tk.StringVar()
var3.set('页脚中间的文字')
fileName = tk.Entry(window, textvariable=var3, width=40).place(x=100, y=140)

tk.Label(window, text='页脚右边的文字').place(x=10, y=200)
var4 = tk.StringVar()
var4.set('页脚右边的文字')
fileName = tk.Entry(window, textvariable=var4, width=40).place(x=100, y=200)


def convertCore(xlsxFiles, path):
    for xlsxFile in xlsxFiles:
        fileDir = path+'/'+xlsxFile

        wb = openpyxl.load_workbook(fileDir)
        for ws in wb.worksheets:
            fillValue = ws.cell(2, 1).value
            if ws.cell(2, 1).value != "":
                if '管线类型：' in fillValue:
                    maxColumnLetter = get_column_letter(ws.max_column)
                    mergeRange = "A2:"+maxColumnLetter+"2"
                    ws.merge_cells(mergeRange)
                    ws.cell(2, 1).value = fillValue
            ws.HeaderFooter.differentOddEven = True
            ws.oddFooter.left = _HeaderFooterPart('探测单位：中国电建集团华东勘测设计研究院有限公司')
            ws.oddFooter.center = _HeaderFooterPart('制表者：段磊仔    校核者：陈皎')

            ws.oddFooter.right.size = 8
            ws.oddFooter.left.size = 8
            ws.oddFooter.center.size = 8
            ws.oddFooter.right.font = "宋体"
            ws.oddFooter.right.text = "日期:2019-5      第 &[Page]-2 页    共 &N-2 页"
            ws.evenFooter.left = _HeaderFooterPart('探测单位：中国电建集团华东勘测设计研究院有限公司')
            ws.evenFooter.center = _HeaderFooterPart('制表者：段磊仔    校核者：陈皎')
            ws.evenFooter.right.size = 8
            ws.evenFooter.left.size = 8
            ws.evenFooter.center.size = 8
            ws.evenFooter.right.font = "宋体"
            ws.evenFooter.right.text = "日期:2019-5      第 &[Page]-2 页    共 &N-2 页"

        dataFrame = pd.read_excel(xlsxFile, sheet_name='目录')
        dataDict = dataFrame.to_dict(orient='list')
        wb.remove(wb['目录'])
        ws00 = wb.create_sheet("目录", 0)
        ws00.row_dimensions[1].height = 15.6
        ws00.row_dimensions[2].height = 50.1
        ws00.row_dimensions[3].height = 51

        ws00.merge_cells("A2:C2")
        ws00.cell(2, 1).value = "目     录"
        ws00.cell(2, 1).alignment = alignObj1
        ws00.cell(2, 1).font = fontObj4
        ws00.column_dimensions['A'].width = 11.38
        ws00.column_dimensions['B'].width = 90.36
        ws00.column_dimensions['C'].width = 9.75

        splitStr = ' ………………………………………………… '
        fillStr = '………………………………………………………………………………………………………………'
        startNum = 4

        for i in dataDict[u'管 线 点 成 果 表']:
            if pd.isnull(i) == False:
                if splitStr in str(i):
                    itemValue = i.split(splitStr)
                    constructorExcel(
                        ws00, startNum, itemValue[0], fillStr, itemValue[1])
                    startNum = startNum+1
        ws1 = wb.create_sheet("封面", 0)

        # 调整列宽
        ws1.column_dimensions['A'].width = 117.88

        # 调整行高
        for i in range(1, 9):
            ws1.row_dimensions[i].height = 60

        ws1['A2'] = "市政文锦渠及东湖公园暗涵综合整治和清污剥离工程—文锦渡口岸泵站"
        ws1['A2'].font = fontObj1
        ws1['A2'].alignment = alignObj1
        ws1['A3'] = "管线点成果表"
        ws1['A3'].font = fontObj2
        ws1['A3'].alignment = alignObj1
        ws1['A8'] = "二〇一九年九月"
        ws1['A8'].font = fontObj3
        ws1['A8'].alignment = alignObj1

        img = Image('images/img.png')
        newsize = (567, 71)
        img.width, img.height = newsize  # 这两个属性分别是对应添加图片的宽高

        img.anchor = 'A7'

        ws1.add_image(img)

        newFileDir = path+'/' + 'new_'+xlsxFile

        wb.save(newFileDir)
        statusStr.set('转换成功')


def batchConvert():
    filePath = var1.get()
    # myPathRoot(filePath)
    if filePath == "":
        statusStr.set('请选择文件夹！！')
        return

    xlsx_file_number = glob.glob(pathname=filePath+'/' + r'*.xlsx')
    print(len(xlsx_file_number))

    if len(xlsx_file_number) == 0:
        statusStr.set('不存在.xlsx的文件')
        return

    xlsxFiles = (fn for fn in os.listdir(filePath) if fn.endswith('.xlsx'))
    convertCore(xlsxFiles, filePath)


tk.Button(window, text='开始脚注', width=50,
          command=batchConvert).place(x=10, y=260)
window.mainloop()
